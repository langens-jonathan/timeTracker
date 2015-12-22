import couchdb
from datetime import date
from openpyxl import Workbook
from openpyxl import utils
from openpyxl import load_workbook
from random import randint

# the name of the log file
logfilename = "timetracker.log"

# the name of the info file
infofilename = "timetracker.info"

# the location of the public file
filePrefix = '../public/'

# an array containing string representations of all the months for output purposes
months = ["january", "february", "march", "april", "may", "june", "juli", "august", "september", "october", "november", "december"]

# convert the date representation of an int to a string
def stringDate(i):
    return months[i]

# the log function prints to the terminal and to a log file
def log(str):
    print str
    with open(logfilename, "a") as logFile:
        logFile.write(str + '\n')
        
# ensures a workbook exists
def ensure(location):
    try:
        with open(location) as file:
            pass
    except IOError as e:
        log("[!] File Error %s, file %s does not exist" % (e, location))
        log("[*] Creating file... %s" % (location))
        wb = Workbook()
        wb.save(location)

# read from the info file
def readInfoFile():
    file = open(infofilename, 'r')
    month = int(file.readline())
    year = int(file.readline())
    return (month, year)

# function that checks if a file exists
def fileExists(filename):
    try:
        with open(filename) as file:
            return True
    except IOError as e:
        return False

# fills the front of a string with zeros...
def fillFrontWithZeros(str, length):
    while(len(str) < length):
        str = '0' + str
    return str

# saves the filename to the database so ember can display it
def saveFileToDatabase(filename):
    tdate = date.today()
    id = 'file_2_' + str(tdate.year) + str(tdate.month) + str(tdate.day)
    idnumber = randint(0, 999999999999999)
    id = id + fillFrontWithZeros(str(idnumber), 16)
    doc = {'_id': id, 'data':{'filename':filename}}
    print("creating doc with id: %s and filename: %s" % (id, filename))
    db.save(doc)
        
# starting the log entry
log("[*] Starting time tracker couchDB > Excell script")
log("[*] Date: %s" % (date.today()))

# setting up the couch db server
serverAddress = 'http://192.168.0.200:5984/'
log("[*] Connecting to database server \'%s\'." % (serverAddress))
couch = couchdb.Server(serverAddress)

# the database we use is timetracker
databasename = 'timetracker'
log("[*] Opening database: \'%s\'." % (databasename))
db = couch[databasename]

# first we get the last month that was fully cleared
# if we find something for this month (or earlier) we just append it to the excel sheet
(lastClearedMonth, lastClearedYear) = readInfoFile()
log("[*] The last month that was fully cleared was: %s %i." % (stringDate(lastClearedMonth), lastClearedYear))


# the 2 arrays that hold all the new and previous activities
# the old activities will be purged from the database, the previous
# file will be overwritten
newActivities = {}
previousActivities = {}
oldActivities = {}

# getting the current and previous month in an integer represenation
currentTime = date.today()
currentMonth = currentTime.month - 1
previousMonth = (currentTime.month + 10) % 12

log("[*] The current month is %s, the previous month is %s." %(stringDate(currentMonth), stringDate(previousMonth)))

# to see if we have to clear out last months data
clearLastMonth = lastClearedMonth <= currentMonth
inString = "have to clear "
if not clearLastMonth:
    inString = "not clear "
log("[*] We will " + inString + "the latest months data.")

# we iterate over all items in our database
for id in db:
    o = db[id]
    
    obj = o['data']
    # check if the item has an start time (this is a poor way to verify that the o is an activity)
    if 'startTime' in obj:
        # we get the start time out of the string in a rather complicated way, the unabridged string is: 2015-12-13T19:41:54.974Z
        # so we take the part from 5 to 7, convert that to an int, an subtract 1 to get a representation that works with our array
        startTime = int(obj['startTime'][5:7]) - 1
        if startTime == currentMonth:
            if not obj['username'] in newActivities:
                newActivities[obj['username']] = []
            newActivities[obj['username']].append(o)
        else:
            if startTime == previousMonth and clearLastMonth:
                if not obj['username'] in previousActivities:
                    previousActivities[obj['username']] = []
                previousActivities[obj['username']].append(o)
            else:
                if not obj['username'] in oldActivities:
                    oldActivities[obj['username']] = {}
                if not startTime in oldActivities[obj['username']]:
                    oldActivities[obj['username']][startTime] = []
                oldActivities[obj['username']][startTime].append(o)
        
log("[*] Writing activity to [username]_current.xlsx")
for user, activities in newActivities.iteritems():
    wbCurrent = Workbook()
    sheet = wbCurrent.active
    sheet.title = "Activity"
    row = 1
    for activity in activities:
        sheet[utils.get_column_letter(1) + str(row)] = activity['data']['description']
        sheet[utils.get_column_letter(2) + str(row)] = activity['data']['startTime']
        sheet[utils.get_column_letter(3) + str(row)] = activity['data']['endTime']
        sheet[utils.get_column_letter(4) + str(row)] = activity['data']['extraInfo']
        row += 1
    if not fileExists(filePrefix + user + '.currentMonth.xlsx'):
        saveFileToDatabase(user + '.currentMonth.xlsx')
    wbCurrent.save(filePrefix + user + '.currentMonth.xlsx')
        
for user, activities in previousActivities.iteritems():
    wbCurrent = Workbook()
    sheet = wbCurrent.active
    sheet.title = "Activity"
    row = 1
    year = activities[0]['data']['startTime'][0:4]
    month = int(activities[0]['data']['startTime'][5:7])
    for activity in activities:
        sheet[utils.get_column_letter(1) + str(row)] = activity['data']['description']
        sheet[utils.get_column_letter(2) + str(row)] = activity['data']['startTime']
        sheet[utils.get_column_letter(3) + str(row)] = activity['data']['endTime']
        sheet[utils.get_column_letter(4) + str(row)] = activity['data']['extraInfo']
        row += 1
        db.delete(activity)
    if not fileExists(filePrefix + user + '.' + year + '.' + str(month) + '.xlsx'):
        saveFileToDatabase(user + '.' + year + '.' + str(month) + '.xlsx')
    wbCurrent.save(user + '.' + year + '.' + str(month) + '.xlsx')

for user, months in oldActivities.iteritems():
    for mon, activities in months.iteritems():
        year = activities[0]['data']['startTime'][0:4]
        month = int(activities[0]['data']['startTime'][5:7])
        filename = user + '.' + year + '.' + str(month) + '.xlsx'
        if not fileExists(filePrefix + filename):
            saveFileToDatabase(filename)
        ensure(filename)
        wbCurrent = load_workbook(filename)
        sheet = wbCurrent.active
        sheet.title = "Activity"
        row = sheet.get_highest_row() + 1
        for activity in activities:
            sheet[utils.get_column_letter(1) + str(row)] = activity['data']['description']
            sheet[utils.get_column_letter(2) + str(row)] = activity['data']['startTime']
            sheet[utils.get_column_letter(3) + str(row)] = activity['data']['endTime']
            sheet[utils.get_column_letter(4) + str(row)] = activity['data']['extraInfo']
            row += 1
            db.delete(activity)
        wbCurrent.save(filePrefix + filename)
